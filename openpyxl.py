from openpyxl import Workbook
from openpyxl.styles import Font
wb=Workbook()
sheet=wb.active
sheet.title="language"
wb.create_sheet(title="capital")
lang=["kannada","telugu","tamil"]
state=["karnataka","telangana","tamilnadu"]
capital=["belangalaru","hyderbad","chennai"]
code=['KA','TS','TN']
sheet.cell(row=1,column=1).value="state"
sheet.cell(row=1,column=2).value="language"
sheet.cell(row=1,column=3).value="code"
ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft
for i in range(2,5):
    sheet.cell(row=i, column=1).value = state[i-2]
    sheet.cell(row=i, column=2).value = lang[i-2]
    sheet.cell(row=i, column=3).value = code[i-2]
wb.save("demo.xlsx")
sheet=wb["capital"]
sheet.cell(row=1,column=1).value="state"
sheet.cell(row=1,column=2).value="capital"
sheet.cell(row=1,column=3).value="code"
ft=Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font=ft
for i in range(2,5):
    sheet.cell(row=i, column=1).value = state[i-2]
    sheet.cell(row=i, column=2).value = capital[i-2]
    sheet.cell(row=i, column=3).value = code[i-2]
wb.save("demo.xlsx")
srchCode=input("enter the capital for the code")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==srchCode:
        print("the corresponding capital code",srchCode,"is",sheet.cell(row=i,column=2).value)
sheet=wb["language"]
srchCode=input("enter the language for the code")
for i in range(2,5):
    data=sheet.cell(row=i,column=3).value
    if data==srchCode:
        print("the corresponding language code",srchCode,"is",sheet.cell(row=i,column=2).value)
wb.close()



